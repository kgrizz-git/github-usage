"""XLSX writer for report exports.

Writes a workbook with one sheet per major report section. Sheet names are
truncated to 31 characters (Excel limit).

Invariants:
- ``_safe_cell`` prefixes values starting with ``=``, ``+``, ``-``, or ``@``
  with a single quote to prevent Excel formula injection.
- ``None`` values are written as empty cells.
- The orchestrator routes stdout output through ``sys.stdout.buffer``
  (binary mode); the writer itself just calls ``wb.save(file_obj)``.
"""

from __future__ import annotations

from collections.abc import Callable

_FORMULA_PREFIXES = ("=", "+", "-", "@")
_MAX_SHEET_NAME = 31

WriteSheetFn = Callable[[str, str, list], None]


def _safe_cell(value):
    """Return ``value`` with Excel formula characters escaped.

    ``None`` becomes an empty string. String values starting with ``=``,
    ``+``, ``-``, or ``@`` are prefixed with a single quote. Numbers and
    other types are passed through unchanged.
    """
    if value is None:
        return ""
    if isinstance(value, str) and value and value[0] in _FORMULA_PREFIXES:
        return "'" + value
    return value


def _truncate_sheet_name(name: str) -> str:
    return name[:_MAX_SHEET_NAME]


def _make_write_sheet(wb):
    """Return a write_sheet closure bound to ``wb``.

    Module-level section helpers take this closure as their first
    argument so they can append rows without constructing their own
    workbook.
    """
    from openpyxl.styles import Font, PatternFill

    def write_sheet(name: str, title: str, rows: list) -> None:
        ws = wb.create_sheet(title=_truncate_sheet_name(name))
        title_font = Font(bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.append([_safe_cell("=" * 50)])
        ws.append([f" {title}"])
        ws.append([_safe_cell("=" * 50)])
        for row in rows:
            ws.append([_safe_cell(cell) for cell in row])
        for cell in ws[2]:
            cell.font = title_font
            cell.fill = title_fill

    return write_sheet


def _write_metadata_sheet(write_sheet: WriteSheetFn, data: dict) -> None:
    write_sheet(
        "Metadata",
        "Report Metadata",
        [
            ["Username", data.get("username", "")],
            ["Period", data.get("period", "")],
            ["Generated At", data.get("generated_at", "")],
        ],
    )


def _write_actions_sheet(write_sheet: WriteSheetFn, data: dict) -> None:
    actions = data.get("actions") or {}
    write_sheet(
        "Actions",
        "Actions Usage",
        [
            [
                "Minutes",
                actions.get("minutes"),
                actions.get("minutes_limit"),
                actions.get("minutes_percent"),
            ],
            [
                "Storage (avg MB)",
                actions.get("storage_avg_mb"),
                actions.get("storage_limit_mb"),
                actions.get("storage_percent"),
            ],
        ],
    )
    sku = actions.get("sku_breakdown") or {}
    if sku:
        sku_rows = [["SKU", "Minutes", "Storage GB-Hrs", "Gross", "Discount", "Net"]]
        for sku_name, sku_data in sku.items():
            sku_rows.append(
                [
                    sku_name,
                    sku_data.get("minutes"),
                    sku_data.get("storage_gb_hours"),
                    sku_data.get("gross"),
                    sku_data.get("discount"),
                    sku_data.get("net"),
                ]
            )
        write_sheet("SKU Breakdown", "Actions SKU Breakdown", sku_rows)


def _write_copilot_sheet(write_sheet: WriteSheetFn, data: dict) -> None:
    copilot = data.get("copilot") or {}
    write_sheet(
        "Copilot",
        "Copilot Usage",
        [
            ["Total Requests", copilot.get("total_requests")],
            ["Total Gross", copilot.get("total_gross")],
            ["Total Discount", copilot.get("total_discount")],
            ["Total Net", copilot.get("total_net")],
        ],
    )
    by_model = copilot.get("by_model") or {}
    if by_model:
        model_rows = [["Model", "Requests", "Gross", "Discount", "Net"]]
        for model_name, model_data in by_model.items():
            if isinstance(model_data, dict):
                model_rows.append(
                    [
                        model_name,
                        model_data.get("requests"),
                        model_data.get("gross"),
                        model_data.get("discount"),
                        model_data.get("net"),
                    ]
                )
        write_sheet("By Model", "Copilot By Model", model_rows)


def _write_git_lfs_sheet(write_sheet: WriteSheetFn, data: dict) -> None:
    git_lfs = data.get("git_lfs") or {}
    write_sheet(
        "Git LFS",
        "Git LFS",
        [
            ["Total Gross", git_lfs.get("total_gross")],
            ["Total Discount", git_lfs.get("total_discount")],
            ["Total Net", git_lfs.get("total_net")],
        ],
    )


def _write_monthly_costs_sheet(write_sheet: WriteSheetFn, data: dict) -> None:
    costs = data.get("monthly_costs") or {}
    cost_rows = [["Category", "Gross", "Discount", "Net"]]
    for category in ("actions", "copilot", "git_lfs", "total"):
        cat = costs.get(category) or {}
        cost_rows.append([category, cat.get("gross"), cat.get("discount"), cat.get("net")])
    write_sheet("Monthly Costs", "Monthly Costs", cost_rows)


def _write_consumers_sheet(write_sheet: WriteSheetFn, data: dict) -> None:
    consumers = data.get("repo_consumers") or {}
    if consumers.get("by_minutes"):
        rows = [["Repo", "Minutes", "Gross", "Storage Avg MB"]]
        for entry in consumers["by_minutes"]:
            rows.append(
                [
                    entry.get("repo"),
                    entry.get("minutes"),
                    entry.get("gross"),
                    entry.get("storage_avg_mb"),
                ]
            )
        write_sheet("Repos Minutes", "Top Repos by Minutes", rows)
    if consumers.get("by_cost"):
        rows = [["Repo", "Minutes", "Gross", "Storage Avg MB"]]
        for entry in consumers["by_cost"]:
            rows.append(
                [
                    entry.get("repo"),
                    entry.get("minutes"),
                    entry.get("gross"),
                    entry.get("storage_avg_mb"),
                ]
            )
        write_sheet("Repos Cost", "Top Repos by Cost", rows)


def _write_artifact_storage_sheet(write_sheet: WriteSheetFn, data: dict) -> None:
    artifacts = data.get("artifact_storage") or {}
    if artifacts.get("top_repos"):
        rows = [["Repo", "Artifact Bytes"]]
        for entry in artifacts["top_repos"]:
            rows.append([entry.get("repo"), entry.get("artifact_bytes")])
        write_sheet("Artifacts", "Artifact Storage", rows)


def _write_release_assets_sheet(write_sheet: WriteSheetFn, data: dict) -> None:
    releases = data.get("release_assets") or {}
    if releases.get("top_repos"):
        rows = [["Repo", "Release Asset Bytes"]]
        for entry in releases["top_repos"]:
            rows.append([entry.get("repo"), entry.get("release_asset_bytes")])
        write_sheet("Releases", "Release Assets", rows)


def _write_insights_sheet(write_sheet: WriteSheetFn, data: dict) -> None:
    insights = data.get("insights") or []
    if insights:
        write_sheet("Insights", "Key Insights", [[finding] for finding in insights])


def _write_errors_sheet(write_sheet: WriteSheetFn, data: dict) -> None:
    errors = data.get("errors") or {}
    if errors:
        write_sheet("Errors", "Unavailable Data", [[k, v] for k, v in errors.items()])


_SECTION_WRITERS = (
    _write_metadata_sheet,
    _write_actions_sheet,
    _write_copilot_sheet,
    _write_git_lfs_sheet,
    _write_monthly_costs_sheet,
    _write_consumers_sheet,
    _write_artifact_storage_sheet,
    _write_release_assets_sheet,
    _write_insights_sheet,
    _write_errors_sheet,
)


def write(data: dict, file_obj) -> None:
    """Write the report data dict as an XLSX workbook to ``file_obj``."""
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_sheet = _make_write_sheet(wb)
    for writer in _SECTION_WRITERS:
        writer(write_sheet, data)
    wb.save(file_obj)
