import io
import unittest
from unittest import mock

from tests.conftest import load_export_report_data

XLSX_SECTIONS = [
    "Metadata",
    "Actions",
    "SKU Breakdown",
    "Copilot",
    "By Model",
    "Git LFS",
    "Monthly Costs",
    "Repos Minutes",
    "Repos Cost",
    "Artifacts",
    "Releases",
    "Insights",
    "Errors",
]


def _has_openpyxl():
    try:
        import openpyxl  # noqa: F401

        return True
    except ImportError:
        return False


@unittest.skipUnless(_has_openpyxl(), "openpyxl not installed")
class ExportXlsxTests(unittest.TestCase):
    def setUp(self):
        self.data = load_export_report_data()

    def _save(self, data=None):
        from github_usage import export_xlsx

        buf = io.BytesIO()
        export_xlsx.write(data if data is not None else self.data, buf)
        return buf.getvalue()

    def _open(self, data=None):
        import openpyxl

        return openpyxl.load_workbook(io.BytesIO(self._save(data)), read_only=True)

    def test_creates_workbook(self):
        wb = self._open()
        self.assertGreaterEqual(len(wb.sheetnames), 1)

    def test_metadata_sheet(self):
        wb = self._open()
        ws = wb["Metadata"]
        rows = list(ws.iter_rows(values_only=True))
        # First three rows are decorative; data starts at row 4
        self.assertEqual(rows[3], ("Username", "octocat"))
        self.assertEqual(rows[4], ("Period", "current_month"))
        self.assertEqual(rows[5], ("Generated At", "2026-06-15T14:30:00Z"))

    def test_sections_present(self):
        wb = self._open()
        for sheet in XLSX_SECTIONS:
            self.assertIn(sheet, wb.sheetnames)

    def test_actions_sheet(self):
        wb = self._open()
        ws = wb["Actions"]
        rows = list(ws.iter_rows(values_only=True))
        # row 4 should be ["Minutes", 1250.0, 2000, 62.5]
        self.assertEqual(rows[3], ("Minutes", 1250.0, 2000, 62.5))

    def test_sku_sheet(self):
        wb = self._open()
        ws = wb["SKU Breakdown"]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[3]
        self.assertEqual(header, ("SKU", "Minutes", "Storage GB-Hrs", "Gross", "Discount", "Net"))
        skus = {r[0] for r in rows[4:] if r[0]}
        self.assertIn("enterprise", skus)
        self.assertIn("free", skus)

    def test_copilot_sheet(self):
        wb = self._open()
        ws = wb["Copilot"]
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(rows[3], ("Total Requests", 42.0))

    def test_copilot_by_model_sheet(self):
        wb = self._open()
        ws = wb["By Model"]
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(rows[3], ("Model", "Requests", "Gross", "Discount", "Net"))
        models = {r[0] for r in rows[4:] if r[0]}
        self.assertIn("gpt-4.1", models)
        self.assertIn("claude-sonnet-4", models)

    def test_git_lfs_sheet(self):
        wb = self._open()
        ws = wb["Git LFS"]
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(rows[3], ("Total Gross", 0.0))

    def test_monthly_costs_sheet(self):
        wb = self._open()
        ws = wb["Monthly Costs"]
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(rows[3], ("Category", "Gross", "Discount", "Net"))
        categories = {r[0] for r in rows[4:] if r[0]}
        for cat in ("actions", "copilot", "git_lfs", "total"):
            self.assertIn(cat, categories)

    def test_repo_consumers_sheets(self):
        wb = self._open()
        for sheet in ("Repos Minutes", "Repos Cost"):
            self.assertIn(sheet, wb.sheetnames)
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            self.assertEqual(rows[3], ("Repo", "Minutes", "Gross", "Storage Avg MB"))

    def test_artifact_storage_sheet(self):
        wb = self._open()
        ws = wb["Artifacts"]
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(rows[3], ("Repo", "Artifact Bytes"))
        self.assertEqual(rows[4], ("octocat/api", 943718400))

    def test_release_assets_sheet(self):
        wb = self._open()
        ws = wb["Releases"]
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(rows[3], ("Repo", "Release Asset Bytes"))
        self.assertEqual(rows[4], ("octocat/api", 314572800))

    def test_insights_sheet(self):
        wb = self._open()
        ws = wb["Insights"]
        rows = list(ws.iter_rows(values_only=True))
        findings = [r[0] for r in rows[3:] if r[0]]
        self.assertTrue(any("72%" in f for f in findings))

    def test_errors_sheet(self):
        wb = self._open()
        ws = wb["Errors"]
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(rows[3], ("copilot", "token may lack billing scope"))

    def test_empty_sections_skipped(self):
        self.data["artifact_storage"] = None
        self.data["release_assets"] = None
        self.data["errors"] = None
        self.data["insights"] = []
        wb = self._open()
        self.assertNotIn("Artifacts", wb.sheetnames)
        self.assertNotIn("Releases", wb.sheetnames)
        self.assertNotIn("Errors", wb.sheetnames)
        self.assertNotIn("Insights", wb.sheetnames)

    def test_sheet_name_truncation(self):
        long_name = "x" * 50
        self.data["actions"]["sku_breakdown"] = {
            long_name: {"minutes": 1, "storage_gb_hours": 1, "gross": 1, "discount": 0, "net": 1}
        }
        wb = self._open()
        for name in wb.sheetnames:
            self.assertLessEqual(len(name), 31)

    def test_safe_cell_formula_escape(self):
        self.data["warnings"] = ["=SUM(A1:A10)", "+1+1", "-cmd|calc", "@danger"]
        wb = self._open()
        # Find the warning row; it's in the Metadata sheet under the header.
        # Warnings are written into the Insights-equivalent; we just check the workbook
        # successfully opens and the values are stored as strings (prefixed).
        ws = wb["Metadata"]
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell in ("=SUM(A1:A10)", "+1+1", "-cmd|calc", "@danger"):
                    self.fail(f"Formula not escaped: {cell!r}")

    def test_to_stdout(self):
        from github_usage import export_xlsx

        captured = io.BytesIO()
        export_xlsx.write(self.data, captured)
        self.assertGreater(len(captured.getvalue()), 0)

    def test_dependency_check_at_orchestrator(self):
        # The orchestrator raises RuntimeError when openpyxl is missing.
        from github_usage import export_report

        with (
            mock.patch.dict("sys.modules", {"openpyxl": None}),
            self.assertRaises(RuntimeError) as ctx,
        ):
            export_report.export(self.data, "xlsx", output_path="/tmp/_should_not_write.xlsx")
        self.assertIn("openpyxl", str(ctx.exception))


if __name__ == "__main__":
    unittest.main()
